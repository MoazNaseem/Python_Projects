#! /usr/bin/python3

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
num = int(input('Enter an int greater than zero: '))
print(type(num))
wb = Workbook()
sheet = wb.get_active_sheet()
for col in range(2, num + 2):
    sheet.cell(row=1, column=col).value = col - 1
for row in range(2, num + 2):
    sheet.cell(column=1, row=row).value = row - 1
for row in range(2, num + 2):
    print('Creating multiplicationTable', row-1)
    for col in range(2, num + 2):
        horizontalLetter = get_column_letter(col)
        vertical = 'A' + str(row)
        horizontal = horizontalLetter.upper() + str(1)
        sheet.cell(row=row, column=col).value = '= {}*{}'.format(vertical, horizontal)
print('Saving...')
wb.save('multiplicationTable.xlsx')
print('Done!')
